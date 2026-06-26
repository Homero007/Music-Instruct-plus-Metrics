"""
Pruebas de las adiciones de la Fase 4 y 5 (KAD, stats, visualizaciones,
entregables, % CLAP). Usan datos sintéticos numpy y backend matplotlib 'Agg';
no requieren modelos preentrenados.

Ejecutar:  python test_phase4_5.py
"""

from __future__ import annotations

import os
import sys
import tempfile
from pathlib import Path

import matplotlib
matplotlib.use("Agg")

import numpy as np

# Permite importar los módulos hermanos (kad_score, stats_tests, ...) tal como
# hace el resto de la etapa new_metrics.
_NM = Path(__file__).resolve().parents[1] / "new_metrics"
if str(_NM) not in sys.path:
    sys.path.insert(0, str(_NM))

import kad_score
import stats_tests
import visualizations
import deliverables


def test_kad_identical_vs_shifted():
    rng = np.random.default_rng(0)
    real = rng.normal(0, 1, (40, 16))
    same = rng.normal(0, 1, (40, 16))      # misma distribución
    shifted = rng.normal(5, 1, (40, 16))   # claramente distinta

    sigma = kad_score.median_bandwidth(real)
    assert sigma > 0

    kad_same = kad_score.mmd2_unbiased(real, same, sigma)
    kad_shift = kad_score.mmd2_unbiased(real, shifted, sigma)
    # Distribuciones iguales -> KAD cercano a 0; desplazada -> claramente mayor.
    assert kad_shift > kad_same
    assert abs(kad_same) < 0.1
    assert kad_shift > 0.2

    core = kad_score.compute_kad_from_embeddings(real, shifted)
    assert core["embedding_dim"] == 16
    assert core["kad"] == kad_shift or abs(core["kad"] - kad_shift) < 1e-9


def test_kruskal_and_dunn():
    rng = np.random.default_rng(1)
    groups = {
        "model_a": list(rng.normal(0.4, 0.05, 30)),
        "model_b": list(rng.normal(0.4, 0.05, 30)),
        "model_c": list(rng.normal(0.7, 0.05, 30)),  # claramente mejor
    }
    result = stats_tests.analyze_metric(groups, metric_name="clap_score")
    assert result["kruskal_wallis"]["significant"] is True
    assert result["posthoc_dunn"] is not None
    comps = {(c["model_a"], c["model_b"]): c for c in result["posthoc_dunn"]["comparisons"]}
    # model_c difiere significativamente de a y de b.
    pair_ca = comps.get(("model_a", "model_c")) or comps.get(("model_c", "model_a"))
    assert pair_ca["significant"] is True
    # a vs b NO deberían diferir.
    pair_ab = comps.get(("model_a", "model_b"))
    assert pair_ab["significant"] is False


def test_per_genre_consistency():
    by_genre = {
        "classical": {"a": [0.3, 0.32], "b": [0.5, 0.52], "c": [0.7, 0.72]},
        "electronic": {"a": [0.31, 0.33], "b": [0.49, 0.51], "c": [0.69, 0.71]},
        "reggaeton": {"a": [0.29, 0.30], "b": [0.50, 0.50], "c": [0.71, 0.73]},
    }
    res = stats_tests.per_genre_consistency(by_genre, higher_is_better=True)
    assert res["kendall_w"] is not None
    assert res["consistent"] is True  # ranking idéntico en los 3 géneros


def test_visualizations_produce_png():
    with tempfile.TemporaryDirectory() as d:
        out = Path(d)
        metrics_by_model = {
            "musicgen_small": {"fad": 1.2, "clap": 0.42, "kld": 0.5, "kad": 0.10},
            "musicgen_medium": {"fad": 0.9, "clap": 0.48, "kld": 0.4, "kad": 0.08},
            "audioldm2": {"fad": 1.5, "clap": 0.38, "kld": 0.6, "kad": 0.15},
        }
        directions = {"fad": False, "clap": True, "kld": False, "kad": False}
        r = visualizations.radar_chart(metrics_by_model, directions, out / "radar.png")
        assert r.exists() and r.stat().st_size > 0

        heat = {
            "musicgen_small": {"classical": 0.4, "electronic": 0.45, "reggaeton": 0.38},
            "audioldm2": {"classical": 0.35, "electronic": 0.40, "reggaeton": 0.30},
        }
        h = visualizations.heatmap_by_genre(heat, out / "heat.png")
        assert h.exists() and h.stat().st_size > 0

        points = {"musicgen_small": (1.2, 0.42), "audioldm2": (1.5, 0.38)}
        s = visualizations.scatter_fad_vs_clap(points, out / "scatter.png")
        assert s.exists() and s.stat().st_size > 0

        groups = {"a": list(np.random.default_rng(0).normal(0.4, 0.05, 30)),
                  "b": list(np.random.default_rng(1).normal(0.5, 0.05, 30))}
        v = visualizations.violin_by_model(groups, out / "violin.png")
        assert v.exists() and v.stat().st_size > 0


def test_deliverables_csv():
    with tempfile.TemporaryDirectory() as d:
        out = Path(d)
        clip_rows = deliverables.build_clip_rows(
            "musicgen_small",
            clip_ids=["c1", "c2"],
            genres={"c1": "classical", "c2": "electronic"},
            tempos={"c1": 90.0, "c2": 128.0},
            clap_scores={"c1": 0.41, "c2": 0.33},
            passt_kld=0.52,
        )
        set_rows = [{
            "model": "musicgen_small", "fad_vggish": 1.2, "fad_pann": 1.1,
            "mean_clap": 0.37, "std_clap": 0.05, "pct_clap_above_025": 80.0,
            "kld": 0.52, "kad": 0.10,
        }]
        paths = deliverables.write_deliverables(clip_rows, set_rows, out)
        metrics_all = Path(paths["metrics_all_csv"])
        set_level = Path(paths["set_level_metrics_csv"])
        assert metrics_all.exists() and set_level.exists()

        head = metrics_all.read_text(encoding="utf-8").splitlines()
        assert head[0] == "model,clip_id,genre,tempo_bpm,clap_score,passt_kld"
        assert "musicgen_small,c1,classical,90.0,0.41,0.52" in head[1]
        set_head = set_level.read_text(encoding="utf-8").splitlines()[0]
        assert set_head == "model,fad_vggish,fad_pann,mean_clap,std_clap,pct_clap_above_025,kld,kad"


def test_evaluate_models_wiring():
    import evaluate_models as em
    em.load_analysis_dependencies()  # carga pd, scipy en globals del módulo
    pd = em.pd
    rng = np.random.default_rng(7)

    # 3 modelos × 3 bloques(géneros) × varias semillas -> scores por bloque.
    rows = []
    blocks = ["classical", "electronic", "reggaeton"]
    model_quality = {"musicgen_small": 0.0, "musicgen_medium": -0.3, "audioldm2": 0.4}
    for model, shift in model_quality.items():
        for block in blocks:
            for _ in range(4):
                rows.append({
                    "block": block,
                    "model": model,
                    "fad": float(1.0 + shift + rng.normal(0, 0.05)),
                    "kad": float(0.10 + 0.1 * shift + rng.normal(0, 0.01)),
                    "clap_score": float(0.45 - shift + rng.normal(0, 0.02)),
                    "tempos_std": float(8.0 + rng.normal(0, 0.5)),
                })
    scores = pd.DataFrame(rows)
    ranking = em.build_ranking(scores)
    model_names = list(model_quality.keys())

    # KAD fluye al ranking junto con FAD/CLAP.
    assert "kad_mean" in ranking.columns

    kruskal = em.run_kruskal_tests(scores, model_names)
    assert not kruskal.empty
    # Con modelos claramente distintos, FAD, KAD y CLAP deben salir significativos.
    assert "kad" in set(kruskal["metric"])
    fad_row = kruskal[kruskal["metric"] == "fad"].iloc[0]
    assert bool(fad_row["significant_alpha_0_05"]) is True
    assert fad_row["dunn_significant_pairs"]  # hay pares significativos

    with tempfile.TemporaryDirectory() as d:
        out = Path(d)
        friedman = em.run_friedman_tests(scores, model_names)
        paths = em.save_outputs(scores, ranking, friedman, out, kruskal=kruskal)
        xlsx = Path(paths["xlsx"])
        assert xlsx.exists() and xlsx.stat().st_size > 0
        # Las 3 gráficas comparativas nuevas se generaron.
        for key in ("radar", "heatmap_clap", "scatter_fad_clap"):
            assert key in paths, f"falta gráfica {key}"
            assert Path(paths[key]).exists()
        # La hoja kruskal_wallis está en el libro.
        from openpyxl import load_workbook
        wb = load_workbook(xlsx)
        assert "kruskal_wallis" in wb.sheetnames


if __name__ == "__main__":
    failures = 0
    for name, fn in sorted(globals().items()):
        if name.startswith("test_") and callable(fn):
            try:
                fn()
                print("ok", name)
            except Exception as exc:  # noqa: BLE001
                failures += 1
                print("FALLO", name, "->", repr(exc))
    if failures:
        sys.exit(f"{failures} prueba(s) fallaron")
    print("TODAS LAS PRUEBAS FASE 4/5 PASARON")
